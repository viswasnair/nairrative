import openpyxl
import csv
import sys

COUNTRY_MAP = {
    'A.G. Riddle': 'USA',
    'Abhijit Banerjee': 'USA',
    'Esther Duflo': 'France',
    'Adolf Hitler': 'Austria',
    'Agatha Christie': 'UK',
    'Agota Kristof': 'Hungary',
    'Alan Moore': 'UK',
    'Dave Gibbons': 'UK',
    'David Lloyd': 'UK',
    'Alastair Reynolds': 'UK',
    'Aldous Huxley': 'UK',
    'Alex Michaelides': 'UK',
    'Amar Chitra Katha': 'India',
    'Amish Tripathi': 'India',
    'Amitav Ghosh': 'India',
    'Amor Towles': 'USA',
    'Andre Agassi': 'USA',
    'Andy Weir': 'USA',
    'Anjali Sachdeva': 'USA',
    'Appupen': 'India',
    'Aravind Adiga': 'India',
    'Art Spiegelman': 'USA',
    'Arthur C. Clarke': 'UK',
    'Gentry Lee': 'USA',
    'Arthur Hailey': 'Canada',
    'Arundhati Roy': 'India',
    'Ashlee Vance': 'USA',
    'Ayn Rand': 'USA',
    'B.R. Ambedkar': 'India',
    'Banu Mushtaq': 'India',
    'Barack Obama': 'USA',
    'Ben Macintyre': 'UK',
    'Bill Bryson': 'USA',
    'Bill Gates': 'USA',
    'Brandon Sanderson': 'USA',
    'Carl Sagan': 'USA',
    'Che Guevara': 'Argentina',
    'Chris Hadfield': 'Canada',
    'Christopher McDougall': 'USA',
    'Christopher Paolini': 'USA',
    'Commando Comics': 'UK',
    'Dan Brown': 'USA',
    'Daniel Kahneman': 'Israel',
    'Daniel Quinn': 'USA',
    'Daniel Tammet': 'UK',
    'David Eagleman': 'USA',
    'David Kushner': 'USA',
    'Deborah Tannen': 'USA',
    'Donella Meadows': 'USA',
    'Douglas Adams': 'UK',
    'Eckhart Tolle': 'Germany',
    'Eileen Magnello': 'UK',
    'Borin Van Loon': 'UK',
    'Elif Shafak': 'Turkey',
    'Eric Van Lustbader': 'USA',
    'Frank Herbert': 'USA',
    'Frank Miller': 'USA',
    'Frederick Forsyth': 'UK',
    'Fyodor Dostoevsky': 'Russia',
    'Gabriel Garcia Marquez': 'Colombia',
    'Gary Chapman': 'USA',
    'Geoff Johns': 'USA',
    'George Orwell': 'UK',
    'Gillian Flynn': 'USA',
    'Grant Snider': 'USA',
    'Greta Thunberg': 'Sweden',
    'H.G. Wells': 'UK',
    'Harper Lee': 'USA',
    'Henry Hazlitt': 'USA',
    'Hermann Hesse': 'Germany',  # birth country; became Swiss citizen later
    'Iain M. Banks': 'UK',
    'Irawati Karve': 'India',
    'Isaac Asimov': 'USA',
    'Isabel Wilkerson': 'USA',
    'J.D. Salinger': 'USA',
    'J.K. Rowling': 'UK',
    'J.R.R. Tolkien': 'UK',
    'Jackie Collins': 'UK',
    'Jacqueline Carey': 'USA',
    'James Clear': 'USA',
    'James Patterson': 'USA',
    'Jared Diamond': 'USA',
    'Jeffrey Archer': 'UK',
    'Jerry Pinto': 'India',
    'Jo Nesbo': 'Norway',
    'John Grisham': 'USA',
    'Jordan Ellenberg': 'USA',
    'Joseph Finder': 'USA',
    'Josy Joseph': 'India',
    'Jules Verne': 'France',
    'Kamala Das': 'India',
    'Ken Follett': 'UK',
    'Kevin Warwick': 'UK',
    'Khaled Hosseini': 'Afghanistan',  # birth country; primary nationality is USA
    'Khushwant Singh': 'India',
    'Kurt Vonnegut': 'USA',
    'Lewis Carroll': 'UK',
    'Lonely Planet': 'Australia',
    'Madeline Miller': 'USA',
    'Mahatma Gandhi': 'India',
    'Malala Yousafzai': 'Pakistan',
    'Manu S. Pillai': 'India',
    'Marcus Aurelius': 'Rome',
    'Marjane Satrapi': 'Iran',
    'Mark Haddon': 'UK',
    'Mark Z. Danielewski': 'USA',
    'Marshall Brain': 'USA',
    'Mary Ann Shaffer': 'USA',
    'Annie Barrows': 'USA',
    'Matt Fitzgerald': 'USA',
    'Matt Wiegle': 'USA',
    'Max Brooks': 'USA',
    'Michael Crichton': 'USA',
    'Michelle Obama': 'USA',
    'Milan Kundera': 'Czech Republic',
    'Mohsin Hamid': 'Pakistan',
    'Naomi Klein': 'Canada',
    'Neal Stephenson': 'USA',
    'Neil Gaiman': 'UK',
    'Nick Lane': 'UK',
    'Orson Scott Card': 'USA',
    'Osamu Tezuka': 'Japan',
    'Ottessa Moshfegh': 'USA',
    'Patrick Olivelle (trans.)': 'Sri Lanka',
    'Patrick Rothfuss': 'USA',
    'Paul Kalanithi': 'USA',
    'Paulo Coelho': 'Brazil',
    'Perumal Murugan': 'India',
    'Peter Thiel': 'USA',
    'Peter Watts': 'Canada',
    'Peter Wohlleben': 'Germany',
    'Philip K. Dick': 'USA',
    'Philip Tetlock': 'USA',
    'Dan Gardner': 'Canada',
    'Prannoy Roy': 'India',
    'Dorab Sopariwala': 'India',
    'Quentin Tarantino': 'USA',
    'R.K. Narayan': 'India',
    'Rabindranath Tagore': 'India',
    'Rachel Carson': 'USA',
    'Rana Ayyub': 'India',
    'Ray Bradbury': 'USA',
    'Raymond Carver': 'USA',
    'Raza Mir': 'India',
    'Rebecca Yarros': 'USA',
    'Richard Dawkins': 'UK',
    'Richard Feynman': 'USA',
    'Richard Thaler': 'USA',
    'Cass Sunstein': 'USA',
    'Robert Arthur': 'USA',
    'Robert Jordan': 'USA',
    'Robert Ludlum': 'USA',
    'Robin Cook': 'USA',
    'Rohinton Mistry': 'India',  # birth country; primary nationality is Canada
    'Rolf Potts': 'USA',
    'S.N. Goenka': 'Myanmar',
    'Salman Rushdie': 'India',  # birth country; primary nationality is UK
    'Sarah J. Maas': 'USA',
    'Scott Adams': 'USA',
    'Scott Lynch': 'USA',
    'Shakespeare': 'UK',
    'Shashi Tharoor': 'India',
    'Shivam Shankar Singh': 'India',
    'Siddhartha Mukherjee': 'India',
    'Sidney Sheldon': 'USA',
    'Steve Brusatte': 'USA',
    'Steven Levitt': 'USA',
    'Stephen Dubner': 'USA',
    'Steven Pinker': 'Canada',
    'Sumit Kumar': 'India',
    'Swati Chaturvedi': 'India',
    'Terry Pratchett': 'UK',
    'Thomas Harris': 'USA',
    'Tom Clancy': 'USA',
    'Tom Rand': 'Canada',
    'Tom Rob Smith': 'UK',
    'Trevor Noah': 'South Africa',
    'Umberto Eco': 'Italy',
    'Unni R.': 'India',
    'V.S. Ramachandran': 'India',
    'Vernor Vinge': 'USA',
    'Vijay Kelkar': 'India',
    'Ajay Shah': 'India',
    'Vikram Seth': 'India',
    'Vladimir Nabokov': 'Russia',
    'Walter Isaacson': 'USA',
    'Will Durant': 'USA',
    'William Dalrymple': 'UK',
    'William Gibson': 'USA',  # birth country; primary nationality is Canada
    'Yann Martel': 'Canada',
    'Yongey Mingyur Rinpoche': 'Nepal',
    'Yuval Noah Harari': 'Israel',
}

# Split compound author strings into individual author names
AUTHOR_SPLIT_MAP = {
    'Abhijit Banerjee & Esther Duflo': ['Abhijit Banerjee', 'Esther Duflo'],
    'Alan Moore & Dave Gibbons': ['Alan Moore', 'David Lloyd'],
    'Arthur C. Clarke & Gentry Lee': ['Arthur C. Clarke', 'Gentry Lee'],
    'Eileen Magnello & Borin Van Loon': ['Eileen Magnello', 'Borin Van Loon'],
    'Mary Ann Shaffer & Annie Barrows': ['Mary Ann Shaffer', 'Annie Barrows'],
    'Philip Tetlock & Dan Gardner': ['Philip Tetlock', 'Dan Gardner'],
    'Prannoy Roy & Dorab Sopariwala': ['Prannoy Roy', 'Dorab Sopariwala'],
    'Richard Thaler & Cass Sunstein': ['Richard Thaler', 'Cass Sunstein'],
    'Shakespeare & Matt Wiegle': ['Shakespeare', 'Matt Wiegle'],
    'Steven Levitt & Stephen Dubner': ['Steven Levitt', 'Stephen Dubner'],
    'Vijay Kelkar & Ajay Shah': ['Vijay Kelkar', 'Ajay Shah'],
}

FORMAT_FIX = {
    'Short Story': 'Short Stories',
    'Illustrated': 'Graphic Novel',
}

USER_ID = '5c8d1748-16ec-45a3-b57e-a8bdb7a7db78'

wb = openpyxl.load_workbook(r'C:\Users\viswa\OneDrive\Docs\reading_list_tmp.xlsx')
ws = wb['Reading List']
headers = [cell.value for cell in ws[1]]

def col(row, name):
    return ws.cell(row, headers.index(name)+1).value

# Build unique authors set (split compound names)
all_authors = {}  # name -> id (1-based)
author_id_counter = [1]

def get_or_create_author(name):
    if name not in all_authors:
        all_authors[name] = author_id_counter[0]
        author_id_counter[0] += 1
    return all_authors[name]

# First pass: collect all authors
for r in range(2, ws.max_row+1):
    raw_author = col(r, 'Author')
    if not raw_author:
        continue
    authors = AUTHOR_SPLIT_MAP.get(raw_author, [raw_author])
    for a in authors:
        get_or_create_author(a)

# Check for missing country mappings
missing = [a for a in all_authors if a not in COUNTRY_MAP]
if missing:
    print('MISSING COUNTRY MAPPINGS:', missing)
    sys.exit(1)

# Second pass: build books and book_authors
books_rows = []
book_authors_rows = []
book_id_counter = [1]

for r in range(2, ws.max_row+1):
    title = col(r, 'Title')
    if not title:
        continue

    raw_author = col(r, 'Author')
    authors = AUTHOR_SPLIT_MAP.get(raw_author, [raw_author])

    year_start = col(r, 'Year Read (Start)')
    year_end = col(r, 'Year Read (End)')

    raw_genre = str(col(r, 'Genre') or '').strip()
    genres = [g.strip() for g in raw_genre.split('/') if g.strip()]
    genre_pg = '{' + ','.join(f'"{g}"' for g in genres) + '}'

    fiction_val = col(r, 'Fiction / Non-Fiction')
    fiction = str(fiction_val).strip().lower() == 'fiction'

    raw_format = str(col(r, 'Format') or '').strip()
    fmt = FORMAT_FIX.get(raw_format, raw_format)

    series = col(r, 'Series') or ''
    raw_series_num = col(r, 'Series #')
    if raw_series_num is None:
        series_num = ''
    else:
        parts = str(raw_series_num).strip().split('-')
        if len(parts) > 1:
            # Range like 1-2 or 1-3: expand to full list
            start, end = int(parts[0]), int(parts[1])
            nums = list(range(start, end + 1))
        else:
            nums = [raw_series_num]
        series_num = '{' + ','.join(str(n) for n in nums) + '}'
    pages = col(r, 'Pages')
    notes = col(r, 'Notes') or ''

    book_id = book_id_counter[0]
    book_id_counter[0] += 1

    books_rows.append({
        'id': book_id,
        'user_id': USER_ID,
        'title': title,
        'year_read_start': year_start,
        'year_read_end': year_end,
        'genre': genre_pg,
        'format': fmt,
        'fiction': fiction,
        'series': series,
        'series_number': series_num if series_num else None,
        'pages': pages,
        'notes': notes,
        'user_added': False,
    })

    for order, author_name in enumerate(authors, start=1):
        book_authors_rows.append({
            'book_id': book_id,
            'author_id': all_authors[author_name],
            'author_order': order,
        })

# Write authors.csv
authors_rows = sorted([
    {'id': aid, 'name': name, 'country': COUNTRY_MAP[name]}
    for name, aid in all_authors.items()
], key=lambda x: x['id'])

base = r'C:\Users\viswa\OneDrive\Docs'

with open(f'{base}\\authors.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['id', 'name', 'country'])
    writer.writeheader()
    writer.writerows(authors_rows)

with open(f'{base}\\books_clean.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['id','user_id','title','year_read_start','year_read_end','genre','format','fiction','series','series_number','pages','notes','user_added'])
    writer.writeheader()
    writer.writerows(books_rows)

with open(f'{base}\\book_authors.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['book_id', 'author_id', 'author_order'])
    writer.writeheader()
    writer.writerows(book_authors_rows)

print(f'authors.csv       — {len(authors_rows)} rows')
print(f'books_clean.csv   — {len(books_rows)} rows')
print(f'book_authors.csv  — {len(book_authors_rows)} rows')
print()
print('Co-authored books:')
for row in book_authors_rows:
    if row['author_order'] > 1:
        book = next(b for b in books_rows if b['id'] == row['book_id'])
        author = next(a for a in authors_rows if a['id'] == row['author_id'])
        print(f"  [{book['id']}] {book['title']} — author {row['author_order']}: {author['name']} ({author['country']})")
